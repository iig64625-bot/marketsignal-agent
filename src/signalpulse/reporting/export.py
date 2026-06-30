"""Export a single report to PDF (fpdf2) and Excel (openpyxl)."""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path


def _read_body(report) -> str:
    if not getattr(report, "markdown_path", None):
        return ""
    try:
        return Path(report.markdown_path).read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def export_to_pdf(report, output_path) -> Path:
    """Render ``report`` to a PDF file (supports CJK via msyh.ttc / simsun.ttc)."""
    from fpdf import FPDF

    body = _read_body(report)
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Pick a CJK font if available; otherwise fall back to Helvetica (no CJK)
    cjk_fonts = [
        (r"C:\Windows\Fonts\msyh.ttc", "msyh", 0),
        (r"C:\Windows\Fonts\msyh.ttf", "msyh", None),
        (r"C:\Windows\Fonts\msyhbd.ttc", "msyhbd", 0),
        (r"C:\Windows\Fonts\simsun.ttc", "simsun", 0),
    ]
    cjk_used = None
    for fpath, family, idx in cjk_fonts:
        if os.path.exists(fpath):
            try:
                if idx is not None:
                    pdf.add_font(family, "", fpath, index=idx)
                else:
                    pdf.add_font(family, "", fpath)
                cjk_used = family
                break
            except Exception:
                continue
    pdf.set_font(cjk_used or "Helvetica", size=10)

    # Header
    pdf.set_font_size(14)
    title = f"SignalPulse Report: {getattr(report, 'title', None) or getattr(report, 'report_type', '?')}"
    pdf.multi_cell(0, 8, title)
    pdf.set_font_size(9)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 5, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)
    pdf.set_font_size(10)

    for line in body.splitlines():
        if line.startswith("# "):
            pdf.set_font_size(13)
            pdf.multi_cell(0, 7, _strip_md(line[2:]))
            pdf.set_font_size(10)
        elif line.startswith("## "):
            pdf.set_font_size(12)
            pdf.multi_cell(0, 6, _strip_md(line[3:]))
            pdf.set_font_size(10)
        elif line.startswith("### "):
            pdf.set_font_size(11)
            pdf.multi_cell(0, 6, _strip_md(line[4:]))
            pdf.set_font_size(10)
        elif line.startswith("- ") or line.startswith("* "):
            pdf.multi_cell(0, 5, "  - " + _strip_md(line[2:]))
        elif line.startswith("---"):
            pdf.ln(2)
        else:
            pdf.multi_cell(0, 5, _strip_md(line))

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(output_path))
    return output_path


def export_to_excel(report, output_path, *, run_id: str | None = None) -> Path:
    """Render ``report`` to an Excel workbook (Report + Signals sheets)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    body = _read_body(report)
    wb = Workbook()

    # Sheet 1: Report text
    ws = wb.active
    ws.title = "Report"
    ws.column_dimensions["A"].width = 110

    ws["A1"] = f"SignalPulse Report: {getattr(report, 'title', None) or getattr(report, 'report_type', '?')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888")

    row = 4
    for line in body.splitlines():
        if line.startswith("# "):
            ws.cell(row=row, column=1, value=_strip_md(line[2:])).font = Font(bold=True, size=13)
        elif line.startswith("## "):
            ws.cell(row=row, column=1, value=_strip_md(line[3:])).font = Font(bold=True, size=12, color="4F46E5")
        elif line.startswith("### "):
            ws.cell(row=row, column=1, value=_strip_md(line[4:])).font = Font(bold=True, size=11)
        else:
            ws.cell(row=row, column=1, value=_strip_md(line))
        row += 1

    # Sheet 2: Signals (if run_id given)
    if run_id:
        try:
            from sqlalchemy import func, select

            from signalpulse.db.session import get_session
            from signalpulse.models.company import Company
            from signalpulse.models.signal import Signal

            ws2 = wb.create_sheet("Signals")
            headers = ["competitor", "signal_type", "count"]
            for col, h in enumerate(headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="4F46E5")
            with get_session() as s:
                rows = s.execute(
                    select(Company.name, Signal.signal_type, func.count(Signal.id))
                    .join(Company, Signal.company_id == Company.id)
                    .where(Signal.crawl_run_id == run_id)
                    .group_by(Company.name, Signal.signal_type)
                ).all()
            for r, (name, stype, count) in enumerate(rows, 2):
                ws2.cell(row=r, column=1, value=name or "?")
                ws2.cell(row=r, column=2, value=stype or "?")
                ws2.cell(row=r, column=3, value=int(count))
            for col_letter in ("A", "B", "C"):
                ws2.column_dimensions[col_letter].width = 22
        except Exception:
            pass  # raw signals are best-effort

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _strip_md(line: str) -> str:
    line = re.sub(r"\*\*(.+?)\*\*", r"\1", line)
    line = re.sub(r"\*(.+?)\*", r"\1", line)
    line = re.sub(r"`([^`]+)`", r"\1", line)
    return line


__all__ = ["export_to_pdf", "export_to_excel"]